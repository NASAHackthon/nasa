#! python3
# dataCollector.py - collect moonquake data and convert it into json file
# data will includes informations like name latitude longtitude depth type time 

import openpyxl, json

# process A type data(mostly deep moonquake)
# replace with datapath in your computor
wb = openpyxl.load_workbook("C:\\Users\\Cooki\\Downloads\\nakamura_2005_dm_locations.xlsx")
sheet = wb['nakamura_2005_dm_locations']
data_list = []
for rowNum in range(2, sheet.max_row):
    data = {}
    name = 'A' + str(sheet.cell(row=rowNum, column=1).value)
    data['name'] = name
    lat = sheet.cell(row=rowNum, column=3).value
    data['latitude'] = lat
    long = sheet.cell(row=rowNum, column=5).value
    data['longtitude'] = long
    dep = sheet.cell(row=rowNum, column=7).value
    data['depth'] = dep
    if (dep):
        if (dep > 700):
            typ = 'Deep Moonquake'
        elif (dep == 0):
            typ = 'Meteoroid Impact'
        else:
            typ = 'Shallow Moonquake'
    else:
        typ = 'Shallow Moonquake'
    data['type'] = typ
    data_list.append(data)
with open("moonquake_data", 'w', encoding="utf-8") as file:
    json.dump(data_list, file, ensure_ascii=False, indent=4)
print("資料已儲存")
